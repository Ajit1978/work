import pandas as pd
from openpyxl import load_workbook

# Read the Excel files
bill_df = pd.read_excel('bill.xlsx', engine='openpyxl')
rr_df = pd.read_excel('rr.xlsx', engine='openpyxl')
ll_df = pd.read_excel('ll.xlsx', engine='openpyxl')

# Create mapping dictionaries
rr_sieve_to_weight = {str(row['Avg MM']).strip(): row['AVG WT'] for _, row in rr_df.iterrows()}
ll_rate_map = dict(zip(ll_df['MM Size'].astype(str), ll_df['ALTR USA Cost/ct']))

# Function to map Sieve/MM to Stone Size
def map_stone_size(sieve_mm):
    sieve_str = str(sieve_mm).strip()
    return rr_sieve_to_weight.get(sieve_str, '')

# Function to map MM Size to Stone Rate
def map_stone_rate(sieve_mm):
    mm_str = str(sieve_mm).strip()
    return ll_rate_map.get(mm_str, '')

# Find the correct column names
sieve_column = [col for col in bill_df.columns if 'Sieve' in col][0]
stone_size_column = [col for col in bill_df.columns if 'Stone Size' in col][0]
stone_rate_column = [col for col in bill_df.columns if 'Stone Rate' in col][0]
stone_wt_total_column = [col for col in bill_df.columns if 'Stone wt. Total' in col][0]
stone_cost_column = [col for col in bill_df.columns if 'Stone Cost' in col][0]
setting_charges_column = [col for col in bill_df.columns if 'Setting Charges' in col][0]

# Apply mappings
bill_df[stone_size_column] = bill_df[sieve_column].apply(map_stone_size)
bill_df[stone_rate_column] = bill_df[sieve_column].apply(map_stone_rate)

# Load the workbook
wb = load_workbook('bill.xlsx')
sheet1 = wb['Sheet1']

# Get the column indices for required columns
n_col_index = list(bill_df.columns).index(stone_size_column) + 1  # Convert to 1-based index
r_col_index = list(bill_df.columns).index(stone_rate_column) + 1  # Convert to 1-based index
stone_wt_total_col_index = list(bill_df.columns).index(stone_wt_total_column) + 1
p_col_index = list(bill_df.columns).index(stone_cost_column) + 1
o_col_index = r_col_index - 1  # Assuming "O" is right before "R"
t_col_index = p_col_index + 2  # Assuming "T" is two columns after "P"
setting_charges_col_index = list(bill_df.columns).index(setting_charges_column) + 1

# Write updated data to the sheet and add formulas for "Stone wt. Total", "Stone Cost", and "Setting Charges"
for row_idx, row in enumerate(bill_df.itertuples(index=False), start=2):  # Start from row 2 (to skip headers)
    for col_idx, value in enumerate(row, start=1):  # Write data column by column
        sheet1.cell(row=row_idx, column=col_idx, value=value)

    # Add formula to "Stone wt. Total" column
    stone_wt_total_formula = f"={chr(64 + n_col_index)}{row_idx}*{chr(64 + r_col_index)}{row_idx}"
    sheet1.cell(row=row_idx, column=stone_wt_total_col_index, value=stone_wt_total_formula)

    # Add formula to "Stone Cost" column
    stone_cost_formula = f"={chr(64 + p_col_index)}{row_idx}*{chr(64 + o_col_index)}{row_idx}"
    sheet1.cell(row=row_idx, column=p_col_index, value=stone_cost_formula)

    # Add formula to "Setting Charges" column
    setting_charges_formula = f"={chr(64 + t_col_index)}{row_idx}*{chr(64 + r_col_index)}{row_idx}"
    sheet1.cell(row=row_idx, column=setting_charges_col_index, value=setting_charges_formula)

# Save the workbook
wb.save('bill_updated.xlsx')

print("Updated Excel file saved as 'bill_updated.xlsx' with formulas for 'Stone wt. Total', 'Stone Cost', and 'Setting Charges' added.")
